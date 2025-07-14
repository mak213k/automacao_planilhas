from openpyxl import Workbook
from openpyxl.drawing.image import Image

wb = Workbook()
ws = wb.active


"""


"""

# Substitua 'nome_do_arquivo.xlsx' pelo caminho do seu arquivo
try:
    workbook = load_workbook('nome_do_arquivo.xlsx')
    # Seleciona a planilha ativa (primeira planilha)
    sheet = workbook.active
    # Ou, se souber o nome da planilha:
    # sheet = workbook['NomeDaPlanilha']

    # Agora você pode acessar as células, linhas e colunas da planilha
    # Exemplo: Imprimir o valor da célula A1
    print(sheet['A1'].value)

    


    # Exemplo: Iterar sobre as linhas e colunas
    for row in sheet.iter_rows():
        for cell in row:

            # create an image
            img = Image('logo.png')

            # add to worksheet and anchor next to cells. 
            #Change 'A1' for letter column and number line

            ws['A1'] = 'imagem'
            ws.add_image(img, 'A1')

            print(cell.value, end=" ")
        print()

except FileNotFoundError:
    print("Erro: Arquivo não encontrado.")
except Exception as e:
    print(f"Ocorreu um erro: {e}")


wb.save('logo.xlsx')