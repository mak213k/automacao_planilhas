import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def encontrar_imagem(desktop, codigo):
    for ext in ['png', 'PNG', 'jpg', 'jpeg', 'gif', 'bmp']:
        caminho = os.path.join(desktop, f"{codigo}.{ext}")
        if os.path.exists(caminho):
            return caminho
    return None

desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
arquivo_excel = os.path.join(desktop_path, "demandas.xlsx")

if not os.path.exists(arquivo_excel):
    print("A planilha 'demandas.xlsx' não foi encontrada na Área de Trabalho.")
    exit()

wb = load_workbook(arquivo_excel)
ws = wb.active

valor = ws["A1"].value
if valor is None:
    print("A célula A1 está vazia. Insira um código e salve a planilha antes de executar.")
    exit()

if isinstance(valor, float):
    codigo = str(int(valor))
else:
    codigo = str(valor).strip()

caminho_imagem = encontrar_imagem(desktop_path, codigo)

if not caminho_imagem:
    print(f"Imagem para o código '{codigo}' não encontrada na Área de Trabalho com extensões comuns.")
    exit()

print(f"Imagem encontrada em: {caminho_imagem}")

img = Image(caminho_imagem)
ws.add_image(img, "B1")

wb.save(arquivo_excel)

print(f"Imagem '{os.path.basename(caminho_imagem)}' inserida com sucesso na célula B1.")