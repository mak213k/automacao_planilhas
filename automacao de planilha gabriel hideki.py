from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

# Caminho do arquivo Excel
arquivo_excel = "agendamentos.xlsx"

# Função para adicionar um agendamento
def add_agendamento(lista):
    while True:
        nome = input("Digite o nome (ou 'sair' para encerrar): ")
        if nome.lower() == 'sair':
            break
        contato = input("Digite o contato: ")
        data = input("Digite a data (DD/MM/AAAA): ")
        hora = input("Digite a hora (HH:MM): ")

        agendamento = [nome, contato, data, hora]
        lista.append(agendamento)
        print("Agendamento adicionado com sucesso!\n")
    return lista

# Função para exibir os agendamentos
def exibir_agendamentos(lista):
    if not lista:
        print("Nenhum agendamento cadastrado.")
        return
    for ag in lista:
        print(f"Nome: {ag[0]} | Contato: {ag[1]} | Data: {ag[2]} | Hora: {ag[3]}")

# Função para salvar na planilha
def salvar_em_planilha(lista):
    if os.path.exists(arquivo_excel):
        wb = load_workbook(arquivo_excel)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Nome", "Contato", "Data", "Hora"])  # Cabeçalhos

    for ag in lista:
        ws.append(ag)

    wb.save(arquivo_excel)
    print(f"Agendamentos salvos em '{arquivo_excel}' com sucesso!")

# Programa principal
if __name__ == "__main__":
    lista_agendamentos = []
    lista_agendamentos = add_agendamento(lista_agendamentos)
    exibir_agendamentos(lista_agendamentos)
    salvar_em_planilha(lista_agendamentos)