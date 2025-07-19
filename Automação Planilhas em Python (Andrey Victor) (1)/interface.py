import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from manipulador import *

def escolher_excel():
    caminho = filedialog.askopenfilename(filetypes=[("Planilhas Excel", "*.xlsx")])
    if not caminho:
        return

    wb = load_workbook(caminho)
    abas = wb.sheetnames

    janela_abas = tk.Toplevel()
    janela_abas.title("Escolha uma aba")

    tk.Label(janela_abas, text="Selecione a aba:").pack(pady=5)
    var_aba = tk.StringVar(value=abas[0])
    menu_abas = tk.OptionMenu(janela_abas, var_aba, *abas)
    menu_abas.pack(pady=10)

    def confirmar_aba():
        dados = ler_planilha_excel(caminho, var_aba.get())
        print(f"\n📄 Dados da aba '{var_aba.get()}':\n")
        for linha in dados:
            print(linha)
        janela_abas.destroy()

    tk.Button(janela_abas, text="Confirmar", command=confirmar_aba).pack(pady=10)

def escolher_arquivo_geral():
    caminho = filedialog.askopenfilename(filetypes=[
        ("Arquivos suportados", "*.csv *.txt")
    ])
    if not caminho:
        return

    if caminho.endswith('.csv'):
        dados = ler_arquivo_csv(caminho)
    elif caminho.endswith('.txt'):
        dados = ler_arquivo_txt(caminho)
    else:
        messagebox.showerror("Erro", "Tipo de arquivo não suportado.")
        return

    criar_planilha(dados, 'convertido_para_excel.xlsx')
    messagebox.showinfo("Sucesso", "Arquivo convertido para Excel com sucesso!")

def iniciar_interface():
    janela = tk.Tk()
    janela.title("Automação de Planilhas")
    janela.geometry("300x250")

    tk.Button(janela, text="Ler Planilha Excel", command=escolher_excel, width=25).pack(pady=15)
    tk.Button(janela, text="Converter CSV/TXT → Excel", command=escolher_arquivo_geral, width=25).pack(pady=15)

    janela.mainloop()
