import csv
from openpyxl import load_workbook, Workbook

def ler_planilha_excel(caminho_arquivo, nome_aba):
    wb = load_workbook(caminho_arquivo)
    planilha = wb[nome_aba]
    dados = [linha for linha in planilha.iter_rows(values_only=True) if any(celula is not None for celula in linha)]
    return dados

def criar_planilha(dados, nome_arquivo='nova_planilha.xlsx'):
    wb = Workbook()
    aba = wb.active
    for linha in dados:
        aba.append(linha)
    wb.save(nome_arquivo)

def ler_arquivo_csv(caminho_arquivo):
    with open(caminho_arquivo, newline='', encoding='utf-8') as f:
        leitor = csv.reader(f)
        return list(leitor)

def ler_arquivo_txt(caminho_arquivo):
    with open(caminho_arquivo, 'r', encoding='utf-8') as f:
        return [linha.strip().split(',') for linha in f]